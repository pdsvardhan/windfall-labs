"""Trendlyne fundamentals ingestion + point-in-time read.

Trendlyne Pro "Data Downloader" exports are SNAPSHOTS (a column-group per file, merged on NSE Code).
A single snapshot is today's fundamentals only — it powers LIVE signals (the DVM screen the owner
trades) but does NOT give historical point-in-time fundamentals. So fundamental features are valid
only on/after their snapshot date; before that they are NaN and fundamental filters correctly fail.
As the owner re-exports over time, snapshots accumulate and history builds forward.
"""
from __future__ import annotations

import datetime as dt

import duckdb
import pandas as pd

from ..config import DATA_DIR
from .store import connect, coverage_summary

FUND_SCHEMA = """
CREATE TABLE IF NOT EXISTS fundamentals (
    ticker VARCHAR NOT NULL,
    snapshot_date DATE NOT NULL,
    reporting_date DATE,
    sector VARCHAR,
    mcap_cr DOUBLE, durability DOUBLE, valuation DOUBLE, momentum_score DOUBLE,
    norm_momentum DOUBLE, pe DOUBLE, fwd_pe DOUBLE, sector_pe DOUBLE, industry_pe DOUBLE,
    pb DOUBLE, eps_ttm DOUBLE, eps_growth DOUBLE, roe DOUBLE, roa DOUBLE, piotroski DOUBLE,
    promoter_holding DOUBLE, promoter_pledge DOUBLE,
    np_qtr_yoy DOUBLE, rev_qtr_yoy DOUBLE, opm DOUBLE,
    rs_nifty_1m DOUBLE, rs_nifty_3m DOUBLE, rs_sector_1m DOUBLE, rs_sector_3m DOUBLE,
    PRIMARY KEY (ticker, snapshot_date)
);
"""

# output column -> Trendlyne header (matched by exact name, first occurrence)
_MAP = {
    "durability": "Trendlyne Durability Score",
    "valuation": "Trendlyne Valuation Score",
    "momentum_score": "Trendlyne Momentum Score",
    "norm_momentum": "Normalized Momentum Score",
    "pe": "PE TTM Price to Earnings",
    "fwd_pe": "Forecaster Estimates 1Y forward PE",
    "sector_pe": "Sector PE TTM",
    "industry_pe": "Industry PE TTM",
    "pb": "Price to Book Value Adjusted",
    "eps_ttm": "Basic EPS TTM",
    "eps_growth": "EPS TTM Growth %",
    "roe": "ROE Annual %",
    "roa": "RoA Annual %",
    "piotroski": "Piotroski Score",
    "mcap_cr": "Market Capitalization",
    "promoter_holding": "Promoter holding latest %",
    "promoter_pledge": "Promoter holding pledge percentage % Qtr",
    "rs_nifty_1m": "Relative returns vs Nifty50 month%",
    "rs_nifty_3m": "Relative returns vs Nifty50 quarter%",
    "rs_sector_1m": "Relative returns vs Sector month%",
    "rs_sector_3m": "Relative returns vs Sector quarter%",
    "np_qtr_yoy": "Net Profit Qtr Growth YoY %",
    "rev_qtr_yoy": "Revenue Growth Qtr YoY %",
    "opm": "Operating Profit Margin Qtr %",
}
NUMERIC_FIELDS = list(_MAP.keys())

# A Trendlyne snapshot older than this is "stale" — the owner should export a fresh one.
# Snapshots are taken roughly monthly; the buffer lets a few days pass before nagging.
SNAPSHOT_STALE_AFTER_DAYS = 35


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "")
    if s in ("", "-", "NA", "N/A", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return pd.to_datetime(str(v)).date()
    except Exception:  # noqa: BLE001
        return None


def _read_sheet(path: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(h).strip() if h is not None else ""
           for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    rows = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[1]]  # NSE Code present
    wb.close()
    return hdr, rows


def parse_trendlyne(paths: list[str]) -> dict[str, dict]:
    """Merge the column-group files on NSE Code -> {ticker: {field: value}}."""
    merged: dict[str, dict] = {}
    for path in paths:
        hdr, rows = _read_sheet(path)
        idx = {name: hdr.index(name) for name in set(_MAP.values()) if name in hdr}
        rdate_i = hdr.index("Result Announced Date") if "Result Announced Date" in hdr else None
        sector_i = hdr.index("sector_name") if "sector_name" in hdr else None
        for r in rows:
            code = str(r[1]).strip().upper()
            ticker = f"{code}.NS"
            rec = merged.setdefault(ticker, {"ticker": ticker})
            for field, name in _MAP.items():
                if name in idx and rec.get(field) is None:
                    rec[field] = _to_float(r[idx[name]])
            if rdate_i is not None and rec.get("reporting_date") is None:
                rec["reporting_date"] = _to_date(r[rdate_i])
            if sector_i is not None and rec.get("sector") is None:
                rec["sector"] = (str(r[sector_i]).strip() if r[sector_i] else None)
    return merged


def ingest(paths: list[str], snapshot_date: str | None = None) -> dict:
    """Parse + store a fundamentals snapshot. Defaults snapshot_date to the latest price bar."""
    con = connect()
    con.execute(FUND_SCHEMA)
    if snapshot_date is None:
        snapshot_date = coverage_summary().get("date_max") or str(dt.date.today())
    snap = dt.date.fromisoformat(str(snapshot_date)[:10])

    recs = parse_trendlyne(paths)
    cols = (["ticker", "snapshot_date", "reporting_date", "sector"] + NUMERIC_FIELDS)
    out_rows = []
    for rec in recs.values():
        row = {c: rec.get(c) for c in cols}
        row["snapshot_date"] = snap
        out_rows.append(row)
    df = pd.DataFrame(out_rows, columns=cols)

    con.execute("DELETE FROM fundamentals WHERE snapshot_date = ?", [snap])
    con.register("fund_in", df)
    con.execute(f"INSERT INTO fundamentals ({','.join(cols)}) SELECT {','.join(cols)} FROM fund_in")
    n_dur = int(df["durability"].notna().sum())
    return {"snapshot_date": str(snap), "stocks": len(df),
            "with_durability": n_dur, "with_pe": int(df["pe"].notna().sum()),
            "fields": NUMERIC_FIELDS}


def snapshots() -> list[str]:
    con = connect()
    con.execute(FUND_SCHEMA)
    return [str(r[0]) for r in
            con.execute("SELECT DISTINCT snapshot_date FROM fundamentals ORDER BY snapshot_date").fetchall()]


def fundamental_panel(field: str, dates: pd.DatetimeIndex, tickers: list[str]) -> pd.DataFrame:
    """Point-in-time panel: at each date, the value from the latest snapshot on/before that date."""
    con = connect()
    con.execute(FUND_SCHEMA)
    df = con.execute(
        f"SELECT ticker, snapshot_date, {field} AS val FROM fundamentals WHERE {field} IS NOT NULL"
    ).fetchdf()
    if df.empty:
        return pd.DataFrame(index=dates, columns=tickers, dtype=float)
    df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    wide = df.pivot_table(index="snapshot_date", columns="ticker", values="val", aggfunc="last")
    wide = wide.reindex(columns=tickers)
    # forward-fill each snapshot across subsequent dates; NaN before the first snapshot.
    full = wide.reindex(wide.index.union(pd.DatetimeIndex(dates))).sort_index().ffill()
    return full.reindex(pd.DatetimeIndex(dates))


def fundamentals_sector_map() -> dict[str, str]:
    con = connect()
    con.execute(FUND_SCHEMA)
    rows = con.execute(
        "SELECT ticker, sector FROM fundamentals WHERE sector IS NOT NULL").fetchall()
    return {t: s for t, s in rows}


def coverage() -> dict:
    con = connect()
    con.execute(FUND_SCHEMA)
    row = con.execute(
        "SELECT COUNT(DISTINCT ticker), COUNT(DISTINCT snapshot_date), MAX(snapshot_date) "
        "FROM fundamentals").fetchone()
    latest = _to_date(row[2]) if row[2] else None
    age_days = (dt.date.today() - latest).days if latest else None
    return {
        "tickers": row[0] or 0,
        "snapshots": row[1] or 0,
        "latest": str(latest) if latest else None,
        "latest_age_days": age_days,
        "stale": bool(age_days is not None and age_days > SNAPSHOT_STALE_AFTER_DAYS),
        "stale_after_days": SNAPSHOT_STALE_AFTER_DAYS,
    }


# ── screener historical fundamentals (point-in-time, lagged) ──────────────────
# The screener.in ingester (windfall.data.screener_fundamentals) writes a STANDALONE DB of annual
# historical fundamentals (~2006->present). Unlike the single Trendlyne snapshot, this gives a real
# point-in-time history, so own-Durability can be backtested over history, not just live-forward.
#
# Anti-look-ahead: an annual fiscal period (period_end) is only public ~3 months after it ends, so a
# value is applied only from period_end + PIT_LAG_DAYS onward (NaN before). 120d is the locked lag
# (covers the SEBI LODR filing window + the long tail of late filers). ONE DOOR: opened READ-ONLY; if
# a scrape holds the writer lock we return None and the caller falls back to the snapshot — a backtest
# never crashes on a locked or missing store.
PIT_LAG_DAYS = 120
SCREENER_DB_PATH = DATA_DIR / "screener_fundamentals.duckdb"

# engine fundamental field -> SQL expression over screener fundamentals_history. Only durability
# inputs are covered; valuation (PE/PB) needs price and is a separate iteration.
_SCREENER_SQL = {
    "roe": "roe",
    "opm": "opm",
    "roa": "(net_profit_owner / NULLIF(total_assets, 0) * 100)",
    # the snapshot's np_qtr_yoy is quarterly; screener history carries only annual growth, so use
    # np_yoy as the historical profit-growth proxy for the same durability input.
    "np_qtr_yoy": "np_yoy",
}
# Raw fundamental fields that gain real history from the screener store (others stay snapshot-only).
SCREENER_HISTORY_FIELDS = frozenset(_SCREENER_SQL)


def _screener_connect():
    """Read-only connection to the standalone screener DB, or None if missing/locked (scrape running)."""
    try:
        return duckdb.connect(str(SCREENER_DB_PATH), read_only=True)
    except Exception:  # noqa: BLE001 — locked by a running scrape, or not yet created
        return None


def screener_history_panel(field: str, dates: pd.DatetimeIndex, tickers: list[str],
                           lag_days: int = PIT_LAG_DAYS) -> pd.DataFrame | None:
    """Point-in-time historical panel for a durability input from the screener store, or None when
    the field isn't covered / there's no data / the DB is unavailable (caller falls back to snapshot).

    The value for fiscal period_end becomes known at period_end + lag_days and is applied from then
    onward (never before) — so there is no look-ahead. Within `dates` the latest known value is
    forward-filled; cells before the first known date stay NaN."""
    expr = _SCREENER_SQL.get(field)
    if expr is None:
        return None
    con = _screener_connect()
    if con is None:
        return None
    try:
        df = con.execute(
            f"SELECT ticker, period_end, {expr} AS val FROM fundamentals_history "
            f"WHERE confidence IN ('high', 'low') AND ({expr}) IS NOT NULL").fetchdf()
    except Exception:  # noqa: BLE001
        return None
    finally:
        con.close()
    if df.empty:
        return None
    # known-from date = fiscal period end + publication lag (the anti-look-ahead control)
    df["known"] = pd.to_datetime(df["period_end"]) + pd.Timedelta(days=lag_days)
    wide = df.pivot_table(index="known", columns="ticker", values="val", aggfunc="last")
    wide = wide.reindex(columns=tickers)
    idx = pd.DatetimeIndex(dates)
    full = wide.reindex(wide.index.union(idx)).sort_index().ffill()
    return full.reindex(idx)


def screener_coverage(lag_days: int = PIT_LAG_DAYS) -> dict:
    """Coverage of the screener historical store: how many names, and from when (lagged) durability
    history is available. `history_from` is the earliest date any lagged annual becomes known."""
    empty = {"available": False, "tickers": 0, "history_from": None,
             "fields": sorted(SCREENER_HISTORY_FIELDS)}
    con = _screener_connect()
    if con is None:
        return empty
    try:
        row = con.execute(
            "SELECT COUNT(DISTINCT ticker), MIN(period_end) FROM fundamentals_history "
            "WHERE confidence IN ('high', 'low')").fetchone()
    except Exception:  # noqa: BLE001
        return empty
    finally:
        con.close()
    if not row or not row[0] or row[1] is None:
        return empty
    history_from = (pd.to_datetime(row[1]) + pd.Timedelta(days=lag_days)).date()
    return {"available": True, "tickers": int(row[0]), "history_from": str(history_from),
            "fields": sorted(SCREENER_HISTORY_FIELDS)}
